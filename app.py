"""
=====================================================================
 KERAS - Kalkulator Estimasi RAB Sipil
 PT PLN Indonesia Power - UBP Saguling | Bidang Pemeliharaan Sipil
 Versi 2.0
=====================================================================
Perubahan utama dari v1:
  1. Database AHSP terpusat (satu sumber kebenaran) + editor harga.
  2. Perbaikan rumus: casing bore pile, wiremesh rigid, satuan aspal,
     plesteran/suling DPT, galian terasering, galian pondasi.
  3. Validasi input otomatis (volume negatif tidak masuk rekap).
  4. Export ke Excel (berformula, bisa diedit) dan PDF.
  5. Penomoran hierarkis (I, I.1, ...) + kolom bobot %.
  6. Kredensial via st.secrets, figure matplotlib ditutup rapi.
=====================================================================
"""

import io
import json
import inspect
import datetime

import streamlit as st
import numpy as np
import pandas as pd
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt

APP_VERSION = "2.0"

# Streamlit >= 1.49 memakai width="stretch"; versi lama memakai use_container_width.
# Helper ini menjaga aplikasi tetap jalan di kedua versi.
_DUKUNG_WIDTH = "width" in inspect.signature(st.button).parameters
LEBAR_PENUH = {"width": "stretch"} if _DUKUNG_WIDTH else {"use_container_width": True}

# =====================================================================
# KONFIGURASI HALAMAN
# =====================================================================
st.set_page_config(page_title="KERAS - Estimator RAB", layout="centered")

st.markdown(
    """
    <style>
        .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
    </style>
    """,
    unsafe_allow_html=True,
)


# =====================================================================
# SISTEM LOGIN
# ---------------------------------------------------------------------
# Kredensial dibaca dari .streamlit/secrets.toml agar tidak ikut
# ter-upload ke repo publik. Format file secrets:
#
#   [auth]
#   username = "sipil.saguling"
#   password = "GantiPasswordIni!"
#
# Jika file secrets belum dibuat, dipakai nilai fallback di bawah.
# =====================================================================
def kredensial():
    try:
        return st.secrets["auth"]["username"], st.secrets["auth"]["password"]
    except Exception:
        return "sipil.saguling", "Sipil2026!"


if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    st.markdown(
        "<h2 style='text-align: center; color: #005c9a;'>🔒 LOGIN AKSES</h2>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p style='text-align: center;'>Kalkulator Estimasi RAB Sipil (KERAS) - PLTA Saguling</p>",
        unsafe_allow_html=True,
    )

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.divider()
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")

        if st.button("🔑 Masuk", **LEBAR_PENUH):
            u_ok, p_ok = kredensial()
            if username == u_ok and password == p_ok:
                st.session_state.logged_in = True
                st.rerun()
            else:
                st.error("❌ Username atau Password salah!")
    st.stop()


# =====================================================================
# DATABASE HARGA SATUAN (AHSP) - SATU SUMBER KEBENARAN
# ---------------------------------------------------------------------
# Format: kode -> (uraian, satuan, harga default)
# Ubah harga di sini ATAU lewat menu "Database Harga Satuan" di aplikasi.
# =====================================================================
DEFAULT_AHSP = {
    # --- Pekerjaan Persiapan ---
    "prep_desain":    ("Pekerjaan Desain Enjiniring",            "LS",    55552544.46),
    "prep_admin":     ("Pekerjaan Administrasi",                 "LS",    15600545.83),
    "prep_sondir":    ("Pekerjaan Sondir",                       "Titik",  2438892.20),
    "prep_survey":    ("Pekerjaan Pengukuran dan Positioning",   "LS",    13549401.09),
    "prep_k3":        ("Penyelenggaraan SMK3 (K3)",              "LS",    43583110.75),
    "prep_mob":       ("Mobilisasi & Demobilisasi",              "LS",     9865747.23),
    "prep_fasilitas": ("Fasilitas Penunjang Pekerjaan",          "LS",    27098802.18),
    # --- Pekerjaan Tanah & Bongkaran ---
    "bongkar_m3":     ("Pembongkaran Struktur Eksisting",        "m³",      380080.60),
    "bongkar_titik":  ("Pembongkaran Struktur per Titik",        "Titik",   380080.60),
    "galian":         ("Pekerjaan Galian Tanah",                 "m³",      174954.45),
    "urugan_tanah":   ("Urugan Tanah Kembali (Backfill)",        "m³",       94351.17),
    "urugan_sirtu":   ("Urugan Sirtu / Material Pilihan",        "m³",      527814.19),
    "pemadatan":      ("Pemadatan Tanah / Badan Jalan",          "m²",       98640.49),
    # --- Pekerjaan Pasangan & Beton ---
    "pas_batu":       ("Pasangan Batu Kali (1:4)",               "m³",      950000.00),
    "plesteran":      ("Plesteran + Acian / Siaran",             "m²",       65000.00),
    "bekisting":      ("Pekerjaan Bekisting",                    "m²",      316349.06),
    "beton_struktur": ("Beton Struktur Bertulang",               "m³",     1723402.09),
    "beton_siklop":   ("Beton Siklop",                           "m³",     1723402.09),
    "lean_concrete":  ("Lean Concrete (K125)",                   "m³",     1598159.55),
    "beton_fs45":     ("Beton FS 45 (Rigid Pavement)",           "m³",     2029697.10),
    "besi":           ("Pembesian / Tulangan",                   "kg",       23662.00),
    "wiremesh_kg":    ("Wiremesh M10 (per kg terpasang)",        "kg",       23662.00),
    "dowel":          ("Pemasangan Dowel Sambungan",             "Titik",    21973.22),
    # --- Pekerjaan Jalan ---
    "lapis_pondasi":  ("Lapis Pondasi Agregat Kelas A",          "m³",      527814.19),
    "aspal_m3":       ("Aspal Hotmix AC-WC (per m³)",            "m³",     2500000.00),
    "aspal_ton":      ("Aspal Hotmix AC-WC (per ton)",           "ton",    1086956.52),
    "tack_coat":      ("Lapis Perekat (Tack Coat)",              "Liter",    15000.00),
    "prime_coat":     ("Lapis Resap Pengikat (Prime Coat)",      "Liter",    15000.00),
    "guard_rail":     ("Pemasangan Guard Rail",                  "m'",     2941333.18),
    # --- Pondasi Dalam ---
    "bor_pile":       ("Pengeboran Bore Pile",                   "m³",      935537.29),
    "casing":         ("Instalasi Temporary Casing",             "m'",      150000.00),
    # --- Drainase & Proteksi Lereng ---
    "suling":         ("Pipa Suling PVC 2\" + Ijuk",             "Titik",    45000.00),
    "perapihan":      ("Perapihan & Pembersihan Permukaan Lereng", "m²",     25000.00),
    "shotcrete":      ("Shotcrete / Facing Beton K-300",         "m³",     2850000.00),
    "wiremesh_m2":    ("Pemasangan Wiremesh M10 (per m²)",       "m²",      115000.00),
    "soil_nail":      ("Soil Nailing D25 Terpasang",             "Titik",  1250000.00),
}

# Catatan asumsi teknis yang dipakai rumus (ditampilkan ke pengguna)
CATATAN_ASUMSI = {
    "aspal_ton": "Konversi m³ ke ton memakai densitas AC-WC padat (default 2,30 t/m³).",
    "wiremesh_kg": "Berat wiremesh M10-150 ≈ 8,22 kg/m² per lapis (Ø10 mm, spasi 150 mm dua arah).",
    "prime_coat": "Harga default disamakan dengan tack coat — mohon disesuaikan dengan AHSP aktual.",
}


# =====================================================================
# INISIALISASI SESSION STATE
# =====================================================================
if "rekap_proyek" not in st.session_state:
    st.session_state.rekap_proyek = []
if "ahsp" not in st.session_state:
    st.session_state.ahsp = {k: v[2] for k, v in DEFAULT_AHSP.items()}
if "ahsp_ver" not in st.session_state:
    st.session_state.ahsp_ver = 0
if "meta" not in st.session_state:
    st.session_state.meta = {
        "paket": "",
        "lokasi": "PLTA Saguling",
        "tahun": str(datetime.date.today().year),
        "nomor": "",
        "penyusun": "",
    }


def H(kode):
    """Ambil harga satuan aktif dari database AHSP."""
    return float(st.session_state.ahsp.get(kode, DEFAULT_AHSP[kode][2]))


def harga_input(kode, wkey, label=None):
    """
    Input harga satuan yang defaultnya mengikuti database AHSP.
    Key widget disisipi versi database supaya nilainya ikut ter-refresh
    setiap kali pengguna mengubah harga di menu Database Harga Satuan.
    """
    uraian, satuan, _ = DEFAULT_AHSP[kode]
    teks = label or f"AHSP {uraian} (Rp/{satuan})"
    return st.number_input(
        teks,
        value=H(kode),
        min_value=0.0,
        step=1000.0,
        format="%.2f",
        key=f"h_{wkey}_v{st.session_state.ahsp_ver}",
    )


# =====================================================================
# HEADER
# =====================================================================
col_head1, col_head2 = st.columns([3, 1])
with col_head1:
    st.markdown("### Aplikasi Estimator RAB")
    st.caption(
        f"Perhitungan teknis volume dan biaya konstruksi terpadu "
        f"**by Pemeliharaan Sipil SGL** · v{APP_VERSION}"
    )
with col_head2:
    if st.button("🚪 Logout", **LEBAR_PENUH):
        st.session_state.logged_in = False
        st.rerun()

st.divider()

# =====================================================================
# IDENTITAS PROYEK (dipakai untuk kop dokumen export)
# =====================================================================
with st.expander("📄 Identitas Dokumen RAB"):
    m = st.session_state.meta
    m["paket"] = st.text_input("Nama Paket Pekerjaan", value=m["paket"])
    c_id1, c_id2 = st.columns(2)
    m["lokasi"] = c_id1.text_input("Lokasi Pekerjaan", value=m["lokasi"])
    m["tahun"] = c_id2.text_input("Tahun Anggaran", value=m["tahun"])
    c_id3, c_id4 = st.columns(2)
    m["nomor"] = c_id3.text_input("Nomor Dokumen", value=m["nomor"])
    m["penyusun"] = c_id4.text_input("Disusun Oleh", value=m["penyusun"])

# =====================================================================
# DATABASE HARGA SATUAN
# =====================================================================
with st.expander("💰 Database Harga Satuan (AHSP)"):
    st.caption(
        "Ubah harga di sini satu kali, seluruh kategori pekerjaan otomatis mengikuti. "
        "Klik **Terapkan** setelah selesai mengedit."
    )
    df_ahsp = pd.DataFrame(
        [
            {"Kode": k, "Uraian": v[0], "Satuan": v[1], "Harga Satuan": H(k)}
            for k, v in DEFAULT_AHSP.items()
        ]
    )
    edited = st.data_editor(
        df_ahsp,
        hide_index=True,
        **LEBAR_PENUH,
        disabled=["Kode", "Uraian", "Satuan"],
        column_config={
            "Harga Satuan": st.column_config.NumberColumn(format="%.2f", min_value=0.0)
        },
        key="ahsp_editor",
    )

    c_db1, c_db2 = st.columns(2)
    if c_db1.button("✅ Terapkan Harga", **LEBAR_PENUH):
        for _, row in edited.iterrows():
            st.session_state.ahsp[row["Kode"]] = float(row["Harga Satuan"])
        st.session_state.ahsp_ver += 1
        st.rerun()
    if c_db2.button("↩️ Reset ke Default", **LEBAR_PENUH):
        st.session_state.ahsp = {k: v[2] for k, v in DEFAULT_AHSP.items()}
        st.session_state.ahsp_ver += 1
        st.rerun()

    with st.container():
        st.caption("**Catatan asumsi teknis:**")
        for kode, catatan in CATATAN_ASUMSI.items():
            st.caption(f"• {DEFAULT_AHSP[kode][0]} — {catatan}")

st.markdown("---")

# =====================================================================
# BLOK 1: TAMBAH PEKERJAAN
# =====================================================================
st.markdown("### ➕ Tambah Pekerjaan")

jenis_bangunan = st.selectbox(
    "Pilih Jenis Pekerjaan:",
    [
        "0. Pekerjaan Persiapan",
        "1. Saluran Air (Batu/Beton/Siklop)",
        "2. Jalan Perkerasan Lentur (Aspal)",
        "3. Jalan Perkerasan Kaku (Rigid)",
        "4. Pondasi Telapak",
        "5. Dinding Penahan Tanah (Stabilisasi Tebing)",
        "6. Pondasi Bore Pile",
        "7. Proteksi Lereng (Shotcrete & Soil Nailing)",
    ],
    key="navigasi_utama",
)

mode_proyek = st.radio(
    "Metode Pelaksanaan:",
    ["Bangunan Baru", "Rehabilitasi Struktur"],
    horizontal=True,
    key="mode_global",
)

with st.expander("⚙️ Pengaturan Keuangan (OAT & PPN)"):
    overhead_pct = st.number_input("Overhead & Profit (%)", value=10.0, step=1.0, key="global_oh")
    ppn_pct = st.number_input("PPN / Pajak (%)", value=11.0, step=1.0, key="global_ppn")

st.markdown("---")

item_to_add = []
peringatan = []
kategori_pekerjaan = jenis_bangunan
fig = None


def tambah(uraian, volume, satuan, harga):
    """
    Masukkan item ke daftar sementara dengan penjagaan nilai tidak wajar.
    Volume <= 0 ditolak agar kesalahan input dimensi tidak diam-diam
    menghasilkan angka negatif di RAB.
    """
    try:
        volume = float(volume)
    except (TypeError, ValueError):
        return
    if volume > 0:
        item_to_add.append([uraian, volume, satuan, float(harga)])
    else:
        peringatan.append(
            f"**{uraian}** tidak dimasukkan — volume hasil hitungan = {volume:,.2f} {satuan}. "
            f"Periksa kembali dimensi yang diisi."
        )


# =====================================================================
# LOGIKA 0. PEKERJAAN PERSIAPAN
# =====================================================================
if jenis_bangunan == "0. Pekerjaan Persiapan":
    st.markdown("**Item Persiapan**")

    urutan_prep = [
        ("prep_desain", "Pekerjaan Desain Enjiniring"),
        ("prep_admin", "Pekerjaan Administrasi"),
        ("prep_survey", "Pekerjaan Pengukuran & Positioning"),
        ("prep_k3", "Penyelenggaraan SMK3 (K3)"),
        ("prep_mob", "Mobilisasi & Demobilisasi Alat"),
        ("prep_fasilitas", "Fasilitas Penunjang Pekerjaan"),
    ]

    dipilih = []
    for kode, label in urutan_prep:
        aktif = st.checkbox(label, value=True, key=f"0_cb_{kode}")
        if aktif:
            harga = harga_input(kode, f"0_{kode}", label=f"Biaya {label} (Rp)")
            dipilih.append((DEFAULT_AHSP[kode][0], 1.0, "LS", harga))

    show_sondir = st.checkbox("Pekerjaan Sondir", value=True, key="0_cb_son")
    if show_sondir:
        vol_sondir = st.number_input("Jumlah Titik Sondir", value=2.0, min_value=0.0, key="0_v_son")
        h_sondir = harga_input("prep_sondir", "0_sondir")
        dipilih.append(("Pekerjaan Sondir", vol_sondir, "Titik", h_sondir))

    for uraian, vol, sat, hrg in dipilih:
        tambah(uraian, vol, sat, hrg)

    fig, ax = plt.subplots(figsize=(4, 2))
    ax.text(
        0.5, 0.5, "Pekerjaan Persiapan & Umum",
        ha="center", va="center", fontsize=12, fontweight="bold", color="gray",
    )
    ax.set_axis_off()

# =====================================================================
# LOGIKA 1. SALURAN AIR
# =====================================================================
elif jenis_bangunan == "1. Saluran Air (Batu/Beton/Siklop)":
    st.markdown("**Material & Lokasi Perbaikan**")
    tipe_saluran = st.radio(
        "Pilih Tipe Struktur:",
        ["Pasangan Batu", "Beton Bertulang", "Beton Siklop"],
        horizontal=True,
        key="1_tipe",
    )

    st.markdown("Pilih Sisi Kerusakan/Pekerjaan:")
    col_s1, col_s2, col_s3 = st.columns(3)
    c_kiri = col_s1.checkbox("Dinding Kiri", True, key="1_ckiri")
    c_lantai = col_s2.checkbox("Lantai Dasar", True, key="1_clantai")
    c_kanan = col_s3.checkbox("Dinding Kanan", True, key="1_ckanan")

    st.markdown("**Dimensi Saluran**")
    l_atas = st.number_input("Lebar Dalam Atas (m)", value=1.2, min_value=0.0, key="1_la")
    l_bawah = st.number_input("Lebar Dalam Bawah (m)", value=0.8, min_value=0.0, key="1_lb")
    tinggi = st.number_input("Tinggi Saluran (m)", value=1.5, min_value=0.0, key="1_t")
    panjang = st.number_input("Panjang Pekerjaan (m)", value=50.0, min_value=0.0, key="1_p")
    t_atas = st.number_input("Tebal Dinding Atas (m)", value=0.25, min_value=0.0, key="1_ta")
    t_bawah = st.number_input("Tebal Dinding Bawah (m)", value=0.40, min_value=0.0, key="1_tb")
    t_dasar = st.number_input("Tebal Lantai Dasar (m)", value=0.30, min_value=0.0, key="1_td")
    ws = st.number_input(
        "Ruang Kerja Galian per Sisi (m)", value=0.30, min_value=0.0,
        help="Kelonggaran galian di luar badan dinding untuk ruang gerak pekerja.",
        key="1_ws",
    )

    sisi_miring = np.sqrt(((l_atas - l_bawah) / 2) ** 2 + tinggi ** 2)
    vol_1_dinding = ((t_atas + t_bawah) / 2) * sisi_miring * panjang
    vol_lantai_m = l_bawah * t_dasar * panjang

    vol_aktif_kiri = vol_1_dinding if c_kiri else 0
    vol_aktif_kanan = vol_1_dinding if c_kanan else 0
    vol_aktif_lantai = vol_lantai_m if c_lantai else 0
    vol_total = vol_aktif_kiri + vol_aktif_kanan + vol_aktif_lantai

    # Luas permukaan dalam (yang benar-benar diplester / dibekisting)
    luas_dinding_dalam = (sisi_miring * panjang if c_kiri else 0) + (
        sisi_miring * panjang if c_kanan else 0
    )
    luas_lantai_dalam = l_bawah * panjang if c_lantai else 0

    st.markdown("**Pekerjaan & AHSP**")
    if mode_proyek != "Bangunan Baru":
        p_bongkar = st.slider("Persen Bongkaran Area Rusak (%)", 0, 100, 100, key="1_sl_bongk")
        if st.checkbox("Pembongkaran Struktur Eksisting", value=True, key="1_cb_bongk"):
            h_bongkar = harga_input("bongkar_m3", "1_bongkar")
            tambah(
                f"Pembongkaran Struktur Eksisting ({p_bongkar}%)",
                vol_total * (p_bongkar / 100), "m³", h_bongkar,
            )

    if st.checkbox("Pekerjaan Galian", value=True, key="1_cb_gal"):
        h_galian = harga_input("galian", "1_galian")
        # PERBAIKAN v2: galian dinding memperhitungkan ruang kerja, dan
        # galian lantai memakai kedalaman lantai + lantai kerja 5 cm.
        vol_gal_kiri = ((t_bawah + ws) * tinggi * panjang) if c_kiri else 0
        vol_gal_kanan = ((t_bawah + ws) * tinggi * panjang) if c_kanan else 0
        vol_gal_lantai = (l_bawah * (t_dasar + 0.05) * panjang) if c_lantai else 0
        tambah("Pekerjaan Galian Saluran", vol_gal_kiri + vol_gal_kanan + vol_gal_lantai, "m³", h_galian)

    if tipe_saluran == "Pasangan Batu":
        if st.checkbox("Pasangan Batu Kali (1:4)", value=True, key="1_cb_batu"):
            h_batu = harga_input("pas_batu", "1_batu")
            tambah("Pasangan Batu Kali (1:4)", vol_total, "m³", h_batu)
        if st.checkbox("Plesteran + Acian (Permukaan Dalam)", value=True, key="1_cb_ples"):
            h_plester = harga_input("plesteran", "1_plester")
            tambah("Plesteran Saluran Sisi Dalam", luas_dinding_dalam + luas_lantai_dalam, "m²", h_plester)

    elif tipe_saluran in ("Beton Bertulang", "Beton Siklop"):
        n_sisi_bek = st.selectbox(
            "Jumlah Sisi Dinding yang Dibekisting",
            [1, 2],
            index=0,
            help="1 = sisi luar dicor langsung ke tanah galian. "
                 "2 = sisi luar juga memakai bekisting (galian lebih lebar).",
            key="1_nsisi",
        )
        if st.checkbox("Pekerjaan Bekisting", value=True, key="1_cb_bek"):
            h_bek = harga_input("bekisting", "1_bekisting")
            tambah("Pekerjaan Bekisting Saluran", luas_dinding_dalam * n_sisi_bek, "m²", h_bek)

        if tipe_saluran == "Beton Bertulang":
            if st.checkbox("Beton Struktur", value=True, key="1_cb_cor"):
                h_cor = harga_input("beton_struktur", "1_beton")
                tambah("Beton Struktur Saluran", vol_total, "m³", h_cor)
            if st.checkbox("Tulangan Utama D16-200", value=True, key="1_cb_besi"):
                r_besi = st.number_input("Rasio Besi (kg/m³)", value=110.0, min_value=0.0, key="1_r_besi")
                h_besi = harga_input("besi", "1_besi")
                tambah("Tulangan Utama D16-200 Saluran", vol_total * r_besi, "kg", h_besi)
        else:
            if st.checkbox("Beton Siklop", value=True, key="1_cb_cor_sik"):
                h_cor = harga_input("beton_siklop", "1_siklop")
                tambah("Beton Siklop Saluran", vol_total, "m³", h_cor)

    # ---- Visualisasi ----
    fig, ax = plt.subplots(figsize=(6, 4))
    x_kiri, x_kanan = -l_bawah / 2, l_bawah / 2
    dx_atas = (l_atas - l_bawah) / 2

    col_kiri = "saddlebrown" if c_kiri else "#e0e0e0"
    col_kanan = "saddlebrown" if c_kanan else "#e0e0e0"
    col_lantai = "saddlebrown" if c_lantai else "#e0e0e0"

    ax.add_patch(plt.Polygon(
        [[x_kiri, 0], [x_kiri - t_bawah, 0], [x_kiri - dx_atas - t_atas, tinggi], [x_kiri - dx_atas, tinggi]],
        color=col_kiri, ec="black", alpha=0.8))
    ax.add_patch(plt.Polygon(
        [[x_kanan, 0], [x_kanan + t_bawah, 0], [x_kanan + dx_atas + t_atas, tinggi], [x_kanan + dx_atas, tinggi]],
        color=col_kanan, ec="black", alpha=0.8))
    ax.add_patch(plt.Polygon(
        [[x_kiri, 0], [x_kanan, 0], [x_kanan, -t_dasar], [x_kiri, -t_dasar]],
        color=col_lantai, ec="black", alpha=0.8))

    ax.text(0, tinggi / 2, f"Ruang Air\nL:{l_atas}m", ha="center", va="center", color="blue", alpha=0.5)
    ax.set_xlim(-l_atas / 2 - max(t_atas, t_bawah) - 0.5, l_atas / 2 + max(t_atas, t_bawah) + 0.5)
    ax.set_ylim(-t_dasar - 0.5, tinggi + 0.5)
    ax.set_aspect("equal")
    ax.set_xlabel("Lebar Saluran (m)")
    ax.set_ylabel("Tinggi/Kedalaman (m)")
    ax.grid(True, linestyle="--", alpha=0.6)

# =====================================================================
# LOGIKA 2. JALAN PERKERASAN LENTUR (ASPAL)
# =====================================================================
elif jenis_bangunan == "2. Jalan Perkerasan Lentur (Aspal)":
    st.markdown("**Dimensi Jalan**")
    lebar = st.number_input("Lebar (m)", value=6.0, min_value=0.0, key="2_l")
    panjang = st.number_input("Panjang (m)", value=1000.0, min_value=0.0, key="2_p")
    t_aspal = st.number_input("Tebal Aspal (m)", value=0.05, min_value=0.0, key="2_tasp")
    t_base = st.number_input("Tebal Agregat Kelas A (m)", value=0.15, min_value=0.0, key="2_tbase")

    luas_jalan = lebar * panjang
    vol_aspal_m3 = luas_jalan * t_aspal

    st.markdown("**Pekerjaan & AHSP**")
    if mode_proyek == "Bangunan Baru":
        if st.checkbox("Pemadatan Tanah / Badan Jalan", value=True, key="2_cb_grad"):
            tambah("Pekerjaan Pemadatan Tanah / Badan Jalan", luas_jalan, "m²",
                   harga_input("pemadatan", "2_pemadatan"))
        if st.checkbox("Lapis Pondasi Agregat Kelas A", value=True, key="2_cb_base"):
            tambah("Pekerjaan Lapis Pondasi Agregat Kelas A", luas_jalan * t_base, "m³",
                   harga_input("lapis_pondasi", "2_lpa"))
        if st.checkbox("Lapis Resap Pengikat (Prime Coat)", value=True, key="2_cb_prime"):
            rate_prime = st.number_input("Takaran Prime Coat (liter/m²)", value=0.80, min_value=0.0, key="2_rprime")
            tambah("Lapis Resap Pengikat (Prime Coat)", luas_jalan * rate_prime, "Liter",
                   harga_input("prime_coat", "2_prime"))
    else:
        p_bongkar = st.slider("Persen Area Dikupas (%)", 0, 100, 100, key="2_sl_bongk")
        if st.checkbox("Pembongkaran & Pengangkutan Bongkaran Jalan", value=True, key="2_cb_mill"):
            tambah(f"Pembongkaran dan Pengangkutan Bongkaran Jalan Eksisting ({p_bongkar}%)",
                   vol_aspal_m3 * (p_bongkar / 100), "m³", harga_input("bongkar_m3", "2_bongkar"))
        if st.checkbox("Lapis Perekat (Tack Coat)", value=True, key="2_cb_tack"):
            rate_tack = st.number_input("Takaran Tack Coat (liter/m²)", value=0.35, min_value=0.0, key="2_rtack")
            tambah("Lapis Perekat (Tack Coat)", luas_jalan * rate_tack, "Liter",
                   harga_input("tack_coat", "2_tack"))

    # PERBAIKAN v2: satuan aspal bisa m³ atau ton (AHSP hotmix umumnya per ton)
    if st.checkbox("Aspal Hotmix AC-WC", value=True, key="2_cb_asp"):
        satuan_aspal = st.radio("Satuan Pembayaran Aspal:", ["ton", "m³"], horizontal=True, key="2_sat_asp")
        if satuan_aspal == "ton":
            densitas = st.number_input("Densitas Aspal Padat (ton/m³)", value=2.30, min_value=0.1, key="2_dens")
            st.caption(f"Volume {vol_aspal_m3:,.2f} m³ × {densitas} t/m³ = **{vol_aspal_m3 * densitas:,.2f} ton**")
            tambah("Aspal Hotmix AC-WC", vol_aspal_m3 * densitas, "ton", harga_input("aspal_ton", "2_aspton"))
        else:
            tambah("Aspal Hotmix AC-WC", vol_aspal_m3, "m³", harga_input("aspal_m3", "2_aspm3"))

    if st.checkbox("Pemasangan Guard Rail", value=False, key="2_cb_gr"):
        panjang_gr = st.number_input("Panjang Guard Rail (m')", value=100.0, min_value=0.0, key="2_p_gr")
        tambah("Pemasangan Guard Rail", panjang_gr, "m'", harga_input("guard_rail", "2_gr"))

    fig, ax = plt.subplots(figsize=(5, 3))
    ax.add_patch(plt.Rectangle((0, 0), lebar, t_aspal, color="black", label="Aspal AC-WC"))
    ax.add_patch(plt.Rectangle((0, -t_base), lebar, t_base, color="orange", alpha=0.4, label="Agregat Kelas A"))
    ax.set_xlim(-1, lebar + 1)
    ax.set_ylim(-t_base - 0.15, t_aspal + 0.15)
    ax.set_aspect("equal")
    ax.set_xlabel("Lebar Jalan (m)")
    ax.set_ylabel("Ketebalan (m)")
    ax.grid(True, linestyle="--", alpha=0.6)
    ax.legend(loc="upper right", fontsize=8)

# =====================================================================
# LOGIKA 3. JALAN PERKERASAN KAKU (RIGID)
# =====================================================================
elif jenis_bangunan == "3. Jalan Perkerasan Kaku (Rigid)":
    st.markdown("**Dimensi Rigid**")
    lebar = st.number_input("Lebar (m)", value=5.0, min_value=0.0, key="3_l")
    panjang = st.number_input("Panjang (m)", value=500.0, min_value=0.0, key="3_p")
    t_rigid = st.number_input("Tebal Rigid (m)", value=0.25, min_value=0.0, key="3_trig")
    t_lc = st.number_input("Tebal Lantai Kerja (m)", value=0.10, min_value=0.0, key="3_tlc")

    luas_jalan = lebar * panjang

    st.markdown("**Pekerjaan & AHSP**")
    if mode_proyek == "Bangunan Baru":
        if st.checkbox("Pemadatan Tanah / Badan Jalan", value=True, key="3_cb_grad"):
            tambah("Pekerjaan Pemadatan Tanah / Badan Jalan", luas_jalan, "m²",
                   harga_input("pemadatan", "3_pemadatan"))
    else:
        p_bongkar = st.slider("Persen Bongkaran (%)", 0, 100, 100, key="3_sl_bongk")
        if st.checkbox("Pembongkaran Jalan Eksisting", value=True, key="3_cb_bongk"):
            tambah(f"Pembongkaran dan Pengangkutan Bongkaran Jalan Eksisting ({p_bongkar}%)",
                   luas_jalan * t_rigid * (p_bongkar / 100), "m³", harga_input("bongkar_m3", "3_bongkar"))

    if st.checkbox("Lean Concrete (K125)", value=True, key="3_cb_lc"):
        tambah("Pekerjaan Lean Concrete (K125)", luas_jalan * t_lc, "m³",
               harga_input("lean_concrete", "3_lc"))
    if st.checkbox("Pekerjaan Bekisting", value=True, key="3_cb_bek"):
        tambah("Pekerjaan Bekisting Tepi Rigid", (t_rigid + t_lc) * panjang * 2, "m²",
               harga_input("bekisting", "3_bek"))
    if st.checkbox("Beton FS 45", value=True, key="3_cb_rig"):
        tambah("Pekerjaan Beton FS 45", luas_jalan * t_rigid, "m³",
               harga_input("beton_fs45", "3_fs45"))

    # PERBAIKAN v2: wiremesh dihitung per kg (luas x berat x lapis x overlap),
    # dowel dipisah sebagai item per titik sambungan.
    if st.checkbox("Wiremesh M10", value=True, key="3_cb_wm"):
        c_wm1, c_wm2 = st.columns(2)
        lapis_wm = c_wm1.number_input("Jumlah Lapis", value=1, min_value=1, step=1, key="3_lapis_wm")
        berat_wm = c_wm2.number_input("Berat Wiremesh (kg/m²)", value=8.22, min_value=0.0,
                                      help="M10-150 ≈ 8,22 kg/m². M10-200 ≈ 6,17 kg/m².", key="3_berat_wm")
        overlap_wm = st.number_input("Faktor Overlap & Waste", value=1.10, min_value=1.0, key="3_ovl_wm")
        tambah(f"Pemasangan Wiremesh M10 ({lapis_wm} Lapis)",
               luas_jalan * berat_wm * lapis_wm * overlap_wm, "kg",
               harga_input("wiremesh_kg", "3_wm"))

    if st.checkbox("Pemasangan Dowel Sambungan", value=True, key="3_cb_dowel"):
        c_dw1, c_dw2 = st.columns(2)
        jarak_joint = c_dw1.number_input("Jarak Antar Sambungan (m)", value=5.0, min_value=0.1, key="3_jjoint")
        jarak_dowel = c_dw2.number_input("Jarak Antar Dowel (m)", value=0.30, min_value=0.05, key="3_jdowel")
        n_joint = np.floor(panjang / jarak_joint)
        n_dowel_per_joint = np.floor(lebar / jarak_dowel)
        total_dowel = n_joint * n_dowel_per_joint
        st.caption(f"*Estimasi {int(n_joint)} sambungan × {int(n_dowel_per_joint)} batang = "
                   f"**{int(total_dowel)} titik dowel***")
        tambah("Pemasangan Dowel Sambungan", total_dowel, "Titik", harga_input("dowel", "3_dowel"))

    if st.checkbox("Pemasangan Guard Rail", value=False, key="3_cb_gr"):
        panjang_gr = st.number_input("Panjang Guard Rail (m')", value=100.0, min_value=0.0, key="3_p_gr")
        tambah("Pemasangan Guard Rail", panjang_gr, "m'", harga_input("guard_rail", "3_gr"))

    fig, ax = plt.subplots(figsize=(5, 3))
    ax.add_patch(plt.Rectangle((0, 0), lebar, t_rigid, color="gray", hatch="//", label="Beton FS 45"))
    ax.add_patch(plt.Rectangle((0, -t_lc), lebar, t_lc, color="orange", alpha=0.4, label="Lean Concrete"))
    ax.set_xlim(-1, lebar + 1)
    ax.set_ylim(-t_lc - 0.2, t_rigid + 0.2)
    ax.set_aspect("equal")
    ax.set_xlabel("Lebar Jalan (m)")
    ax.set_ylabel("Ketebalan (m)")
    ax.grid(True, linestyle="--", alpha=0.6)
    ax.legend(loc="upper right", fontsize=8)

# =====================================================================
# LOGIKA 4. PONDASI TELAPAK
# =====================================================================
elif jenis_bangunan == "4. Pondasi Telapak":
    st.markdown("**Dimensi Pondasi**")
    p = st.number_input("Panjang Plat (m)", value=1.5, min_value=0.0, key="4_p")
    l = st.number_input("Lebar Plat (m)", value=1.5, min_value=0.0, key="4_l")
    t = st.number_input("Tebal Plat (m)", value=0.3, min_value=0.0, key="4_t")
    jml = st.number_input("Jumlah Titik", value=10, min_value=0, step=1, key="4_jml")

    t_lc = st.number_input("Tebal Lantai Kerja (m)", value=0.05, min_value=0.0, key="4_tlc")
    ws = st.number_input("Ruang Kerja Galian per Sisi (m)", value=0.20, min_value=0.0, key="4_ws")
    h_gali = st.number_input(
        "Kedalaman Galian (m)", value=1.00, min_value=0.0,
        help="Dari muka tanah asli sampai dasar lantai kerja.", key="4_hgali",
    )

    vol_beton = p * l * t * jml
    vol_lc = p * l * t_lc * jml
    vol_galian = (p + 2 * ws) * (l + 2 * ws) * h_gali * jml
    vol_urugan = max(0.0, vol_galian - vol_beton - vol_lc)

    st.markdown("**Pekerjaan & AHSP**")
    if mode_proyek != "Bangunan Baru":
        p_bongkar = st.slider("Persen Bongkaran (%)", 0, 100, 100, key="4_sl_bongk")
        if st.checkbox("Pembongkaran Struktur Eksisting", value=True, key="4_cb_bongk"):
            tambah(f"Pembongkaran Struktur Eksisting ({p_bongkar}%)",
                   vol_beton * (p_bongkar / 100), "m³", harga_input("bongkar_m3", "4_bongkar"))

    if st.checkbox("Pekerjaan Galian", value=True, key="4_cb_gal"):
        tambah("Pekerjaan Galian Pondasi", vol_galian, "m³", harga_input("galian", "4_galian"))
    if st.checkbox("Lean Concrete (K125)", value=True, key="4_cb_lc"):
        tambah("Pekerjaan Lean Concrete (K125)", vol_lc, "m³", harga_input("lean_concrete", "4_lc"))
    if st.checkbox("Pekerjaan Bekisting", value=True, key="4_cb_bek"):
        tambah("Pekerjaan Bekisting Plat Pondasi", (p + l) * 2 * t * jml, "m²",
               harga_input("bekisting", "4_bek"))
    if st.checkbox("Beton Struktur", value=True, key="4_cb_cor"):
        tambah("Beton Plat Pondasi", vol_beton, "m³", harga_input("beton_struktur", "4_beton"))
    if st.checkbox("Tulangan Utama D16-200", value=True, key="4_cb_besi"):
        r_besi = st.number_input("Rasio Besi (kg/m³)", value=150.0, min_value=0.0, key="4_r_besi")
        tambah("Tulangan Utama D16-200 Pondasi", vol_beton * r_besi, "kg", harga_input("besi", "4_besi"))
    if st.checkbox("Urugan Tanah Kembali", value=True, key="4_cb_urug"):
        st.caption(f"*Galian {vol_galian:,.2f} m³ − beton {vol_beton:,.2f} m³ − lantai kerja "
                   f"{vol_lc:,.2f} m³ = **{vol_urugan:,.2f} m³***")
        tambah("Urugan Tanah Kembali Pondasi", vol_urugan, "m³", harga_input("urugan_tanah", "4_urug"))

    fig, ax = plt.subplots(figsize=(5, 3.5))
    ax.add_patch(plt.Rectangle((-(p + 2 * ws) / 2, -h_gali), p + 2 * ws, h_gali,
                               color="saddlebrown", alpha=0.15, label="Area Galian"))
    ax.add_patch(plt.Rectangle((-p / 2, -h_gali + t_lc), p, t, color="gray", ec="black", label="Plat Beton"))
    ax.add_patch(plt.Rectangle((-p / 2, -h_gali), p, t_lc, color="orange", alpha=0.5, label="Lantai Kerja"))
    ax.plot([-(p + 2 * ws) / 2 - 0.5, (p + 2 * ws) / 2 + 0.5], [0, 0], color="saddlebrown", lw=3)
    ax.set_xlim(-(p + 2 * ws) / 2 - 0.6, (p + 2 * ws) / 2 + 0.6)
    ax.set_ylim(-h_gali - 0.3, 0.4)
    ax.set_aspect("equal")
    ax.set_xlabel("Lebar Galian / Pondasi (m)")
    ax.set_ylabel("Kedalaman (m)")
    ax.grid(True, linestyle="--", alpha=0.6)
    ax.legend(loc="lower right", fontsize=8)

# =====================================================================
# LOGIKA 5. DINDING PENAHAN TANAH
# =====================================================================
elif jenis_bangunan == "5. Dinding Penahan Tanah (Stabilisasi Tebing)":
    st.markdown("**Tipe Struktur & Dimensi**")
    tipe_dpt = st.radio(
        "Pilih Tipe Struktur DPT:",
        [
            "Pasangan Batu (Gravity Wall)",
            "Pasangan Batu Bertingkat (Terasering)",
            "Beton Siklop (Gravity Wall)",
            "Beton Siklop Bertingkat (Terasering)",
            "Beton Bertulang (Cantilever)",
        ],
        key="5_tipe",
    )

    panjang = st.number_input("Panjang Total DPT (m)", value=50.0, min_value=0.0, key="5_p")

    def input_suling(luas_muka, prefix):
        """Hitung jumlah titik suling berdasarkan grid pemasangan, bukan asumsi tetap."""
        if not st.checkbox('Pipa Suling-Suling PVC 2" + Ijuk', value=True, key=f"{prefix}_cb_suling"):
            return
        c1, c2 = st.columns(2)
        jh = c1.number_input("Jarak Horizontal Suling (m)", value=2.0, min_value=0.1, key=f"{prefix}_jh_sul")
        jv = c2.number_input("Jarak Vertikal Suling (m)", value=2.0, min_value=0.1, key=f"{prefix}_jv_sul")
        n_suling = np.ceil(luas_muka / (jh * jv)) if luas_muka > 0 else 0
        st.caption(f"*Luas muka depan {luas_muka:,.2f} m² ÷ grid {jh}×{jv} m = **{int(n_suling)} titik***")
        tambah('Instalasi Pipa Suling PVC 2" + Ijuk', n_suling, "Titik",
               harga_input("suling", f"{prefix}_suling"))

    # ------------------------------------------------------------------
    if tipe_dpt in ["Pasangan Batu (Gravity Wall)", "Beton Siklop (Gravity Wall)"]:
        is_siklop = "Siklop" in tipe_dpt
        st.markdown("**Dimensi DPT Gravity Wall**")
        h = st.number_input("Tinggi Dinding (m)", value=4.0, min_value=0.0, key="5_g_h")
        l_bawah = st.number_input("Lebar Dasar/Bawah (m)", value=1.5, min_value=0.0, key="5_g_lb")
        l_atas = st.number_input("Lebar Atas (m)", value=0.4, min_value=0.0, key="5_g_la")
        offset_depan = st.number_input(
            "Kemiringan Sisi Depan (m)", value=0.3, min_value=0.0,
            help="Jarak horizontal kemiringan dari ujung bawah ke ujung atas sisi depan.",
            key="5_g_off",
        )
        ws = st.number_input("Ruang Kerja Galian (m)", value=0.50, min_value=0.0, key="5_g_ws")

        vol_material = ((l_atas + l_bawah) / 2) * h * panjang
        sisi_miring_depan = np.sqrt(h ** 2 + offset_depan ** 2)
        offset_belakang = l_bawah - offset_depan - l_atas
        sisi_miring_belakang = np.sqrt(h ** 2 + max(offset_belakang, 0) ** 2)

        # PERBAIKAN v2: muka depan dan muka belakang dipisah. Sisi belakang
        # menempel tanah/backfill sehingga tidak diplester secara default.
        luas_muka_depan = sisi_miring_depan * panjang
        luas_muka_belakang = sisi_miring_belakang * panjang
        luas_puncak = l_atas * panjang
        vol_galian = (l_bawah + ws) * h * panjang

        if offset_belakang < 0:
            st.warning(
                "⚠️ Lebar atas + kemiringan depan melebihi lebar dasar. "
                "Profil dinding menjadi tidak wajar — periksa kembali dimensinya."
            )

        st.markdown("**Pekerjaan & AHSP**")
        if mode_proyek != "Bangunan Baru":
            p_bongkar = st.slider("Persen Bongkaran (%)", 0, 100, 100, key="5_g_sl_bongk")
            if st.checkbox("Pembongkaran Struktur Eksisting", value=True, key="5_g_cb_bongk"):
                tambah(f"Pembongkaran Struktur Eksisting ({p_bongkar}%)",
                       vol_material * (p_bongkar / 100), "m³", harga_input("bongkar_m3", "5_g_bongkar"))

        if st.checkbox("Pekerjaan Galian", value=True, key="5_g_cb_gal"):
            tambah("Pekerjaan Galian Tebing", vol_galian, "m³", harga_input("galian", "5_g_galian"))

        if not is_siklop:
            if st.checkbox("Pasangan Batu Kali (1:4)", value=True, key="5_g_cb_mat"):
                tambah("Pasangan Batu Kali (1:4)", vol_material, "m³", harga_input("pas_batu", "5_g_mat"))
            if st.checkbox("Plesteran & Siaran DPT", value=True, key="5_g_cb_ples"):
                sisi_ples = st.selectbox(
                    "Permukaan yang Diplester",
                    ["Sisi depan + puncak", "Sisi depan saja", "Depan + belakang + puncak"],
                    key="5_g_sisi_ples",
                )
                if sisi_ples == "Sisi depan saja":
                    luas_ples = luas_muka_depan
                elif sisi_ples == "Sisi depan + puncak":
                    luas_ples = luas_muka_depan + luas_puncak
                else:
                    luas_ples = luas_muka_depan + luas_muka_belakang + luas_puncak
                tambah("Plesteran & Siaran Permukaan DPT", luas_ples, "m²",
                       harga_input("plesteran", "5_g_ples"))
        else:
            if st.checkbox("Pekerjaan Bekisting", value=True, key="5_g_cb_bek"):
                sisi_bek = st.selectbox("Jumlah Sisi Dibekisting", [1, 2], index=1, key="5_g_nsisi")
                luas_bek = luas_muka_depan if sisi_bek == 1 else luas_muka_depan + luas_muka_belakang
                tambah("Pekerjaan Bekisting DPT Siklop", luas_bek, "m²", harga_input("bekisting", "5_g_bek"))
            if st.checkbox("Beton Siklop", value=True, key="5_g_cb_mat_s"):
                tambah("Beton Siklop DPT", vol_material, "m³", harga_input("beton_siklop", "5_g_siklop"))

        if st.checkbox("Urugan Kembali (Backfill)", value=True, key="5_g_cb_timb"):
            vol_timbunan = (0.5 * offset_belakang * h * panjang) if offset_belakang > 0 else 0
            tambah("Pekerjaan Urugan Kembali (Backfill)", vol_timbunan, "m³",
                   harga_input("urugan_tanah", "5_g_urug"))

        input_suling(luas_muka_depan, "5_g")

        # ---- Visualisasi ----
        fig, ax = plt.subplots(figsize=(6, 5))
        ax.add_patch(plt.Polygon(
            [[0, 0], [l_bawah, 0], [offset_depan + l_atas, h], [offset_depan, h]],
            color="#8b9ea8" if is_siklop else "slategray", ec="black", alpha=0.9))
        lebar_timbunan = max(1.0, offset_belakang + 0.5)
        ax.add_patch(plt.Polygon(
            [[l_bawah, 0], [l_bawah + lebar_timbunan, 0],
             [l_bawah + lebar_timbunan, h], [offset_depan + l_atas, h]],
            color="saddlebrown", alpha=0.3, label="Timbunan Tebing"))
        ax.plot([-0.5, l_bawah + lebar_timbunan], [0, 0], color="saddlebrown", lw=3)
        ax.text(l_bawah / 2, 0.2, f"{l_bawah}m", ha="center", va="bottom", fontsize=9, color="white")
        ax.text(offset_depan + l_atas / 2, h - 0.3, f"{l_atas}m", ha="center", va="top", fontsize=9, color="white")
        ax.set_xlim(-1.0, l_bawah + lebar_timbunan + 0.5)
        ax.set_ylim(-0.5, h + 1.0)
        ax.set_aspect("equal")
        ax.set_xlabel("Lebar Struktur (m)")
        ax.set_ylabel("Tinggi/Elevasi (m)")
        ax.grid(True, linestyle="--", alpha=0.6)
        ax.legend(loc="upper right")

    # ------------------------------------------------------------------
    elif tipe_dpt in ["Pasangan Batu Bertingkat (Terasering)", "Beton Siklop Bertingkat (Terasering)"]:
        is_siklop = "Siklop" in tipe_dpt
        jml_tingkat = st.number_input("Jumlah Tingkat (Trap)", value=3, step=1, min_value=1, key="5_ter_jml")
        h_trap = st.number_input("Tinggi per Tingkat (m)", value=2.0, min_value=0.0, key="5_ter_h")
        l_atas = st.number_input("Lebar Atas per Tingkat (m)", value=0.4, min_value=0.0, key="5_ter_la")
        l_bawah = st.number_input("Lebar Bawah per Tingkat (m)", value=1.0, min_value=0.0, key="5_ter_lb")
        l_berm = st.number_input("Lebar Pijakan/Berm antar Tingkat (m)", value=0.5, min_value=0.0, key="5_ter_berm")
        ws = st.number_input("Ruang Kerja Galian per Trap (m)", value=0.50, min_value=0.0, key="5_ter_ws")

        vol_per_trap = ((l_atas + l_bawah) / 2) * h_trap * panjang
        vol_total_mat = vol_per_trap * jml_tingkat
        sisi_miring = np.sqrt(h_trap ** 2 + (l_bawah - l_atas) ** 2)
        luas_muka_depan = sisi_miring * panjang * jml_tingkat
        luas_berm = l_berm * panjang * max(jml_tingkat - 1, 0)

        # PERBAIKAN v2: galian bertingkat dimodelkan sebagai penjumlahan prisma
        # per trap (skala linear terhadap jumlah trap), bukan satu balok penuh
        # setinggi total tebing yang membuat volume membengkak kuadratik.
        vol_galian = (l_bawah + l_berm + ws) * h_trap * panjang * jml_tingkat

        st.info(
            f"Total tinggi tebing tertangani: **{h_trap * jml_tingkat:,.2f} m** · "
            f"Volume struktur: **{vol_total_mat:,.2f} m³**"
        )
        st.caption(
            "*Galian terasering dihitung sebagai prisma bertingkat "
            "(lebar bawah + berm + ruang kerja) × tinggi trap × panjang × jumlah trap. "
            "Untuk tebing dengan kemiringan asli yang landai, verifikasi ulang dengan potongan melintang aktual.*"
        )

        st.markdown("**Pekerjaan & AHSP**")
        if mode_proyek != "Bangunan Baru":
            p_bongkar = st.slider("Persen Bongkaran (%)", 0, 100, 100, key="5_ter_sl_bongk")
            if st.checkbox("Pembongkaran Struktur Eksisting", value=True, key="5_ter_cb_bongk"):
                tambah(f"Pembongkaran Struktur Eksisting ({p_bongkar}%)",
                       vol_total_mat * (p_bongkar / 100), "m³", harga_input("bongkar_m3", "5_ter_bongkar"))

        if st.checkbox("Pekerjaan Galian", value=True, key="5_ter_cb_gal"):
            tambah("Pekerjaan Galian Tebing (Terasering)", vol_galian, "m³",
                   harga_input("galian", "5_ter_galian"))

        if not is_siklop:
            if st.checkbox("Pasangan Batu Kali (1:4)", value=True, key="5_ter_cb_mat"):
                tambah("Pasangan Batu Kali (Terasering)", vol_total_mat, "m³",
                       harga_input("pas_batu", "5_ter_mat"))
            if st.checkbox("Plesteran & Siaran DPT", value=True, key="5_ter_cb_ples"):
                ikut_berm = st.checkbox("Termasuk perkerasan permukaan berm", value=True, key="5_ter_berm_ples")
                tambah("Plesteran & Siaran Permukaan Terasering",
                       luas_muka_depan + (luas_berm if ikut_berm else 0), "m²",
                       harga_input("plesteran", "5_ter_ples"))
        else:
            if st.checkbox("Pekerjaan Bekisting", value=True, key="5_ter_cb_bek"):
                tambah("Pekerjaan Bekisting Terasering Siklop", luas_muka_depan, "m²",
                       harga_input("bekisting", "5_ter_bek"))
            if st.checkbox("Beton Siklop", value=True, key="5_ter_cb_mat_s"):
                tambah("Beton Siklop Terasering", vol_total_mat, "m³",
                       harga_input("beton_siklop", "5_ter_siklop"))

        input_suling(luas_muka_depan, "5_ter")

        # ---- Visualisasi ----
        fig, ax = plt.subplots(figsize=(6, 5))
        x_heel, y_bottom = 0.0, 0.0
        max_x, min_x = 0.0, 0.0
        soil_pts = [[0, 0]]
        for i in range(int(jml_tingkat)):
            x_toe = x_heel + l_bawah
            pts = np.array([
                [x_heel, y_bottom], [x_toe, y_bottom],
                [x_heel + l_atas, y_bottom + h_trap], [x_heel, y_bottom + h_trap],
            ])
            ax.add_patch(plt.Polygon(pts, color="#8b9ea8" if is_siklop else "slategray",
                                     alpha=0.9, ec="black", lw=1.5))
            ax.text(x_heel + l_bawah / 2, y_bottom + 0.1, f"{l_bawah}m",
                    ha="center", va="bottom", fontsize=8, color="white")
            soil_pts.append([x_heel, y_bottom])
            soil_pts.append([x_heel, y_bottom + h_trap])
            if i < jml_tingkat - 1:
                next_x_toe = x_heel - l_berm
                soil_pts.append([next_x_toe, y_bottom + h_trap])
                ax.text(x_heel - l_berm / 2, y_bottom + h_trap + 0.1, f"Berm {l_berm}m",
                        ha="center", va="bottom", fontsize=8, color="saddlebrown")
                x_heel = next_x_toe - l_bawah
            y_bottom += h_trap
            min_x = min(min_x, x_heel)
            max_x = max(max_x, x_toe)

        soil_pts.append([min_x - 2, y_bottom])
        soil_pts.append([min_x - 2, 0])
        ax.add_patch(plt.Polygon(soil_pts, color="saddlebrown", alpha=0.2))
        x_s, y_s = zip(*soil_pts[:-2])
        ax.plot(x_s, y_s, color="saddlebrown", lw=3, label="Tanah / Tebing")
        ax.set_xlim(min_x - 1.5, max_x + 1.5)
        ax.set_ylim(-1, y_bottom + 1.5)
        ax.set_aspect("equal")
        ax.set_xlabel("Jarak Horizontal (m)")
        ax.set_ylabel("Tinggi Elevasi (m)")
        ax.grid(True, linestyle="--", alpha=0.6)
        ax.legend(loc="upper left")

    # ------------------------------------------------------------------
    else:  # Beton Bertulang (Cantilever)
        st.markdown("**Dimensi Dinding Cantilever & Tapak**")
        h = st.number_input("Tinggi Dinding (Stem) (m)", value=4.0, min_value=0.0, key="5_c_h")
        l_base = st.number_input("Lebar Total Base/Tapak (m)", value=2.5, min_value=0.0, key="5_c_lb")
        t_base = st.number_input("Tebal Base/Tapak (m)", value=0.4, min_value=0.0, key="5_c_tb")
        l_toe = st.number_input("Jarak Ujung Depan ke Dinding (Toe) (m)", value=0.5, min_value=0.0, key="5_c_ltoe")
        t_bawah = st.number_input("Tebal Dinding Bawah (m)", value=0.5, min_value=0.0, key="5_c_tbwh")
        t_atas = st.number_input("Tebal Dinding Atas (m)", value=0.3, min_value=0.0, key="5_c_tats")

        l_heel = l_base - l_toe - t_bawah
        if l_heel < 0:
            st.error(
                f"❌ Dimensi tidak konsisten: lebar tapak ({l_base} m) lebih kecil dari "
                f"toe + tebal dinding ({l_toe + t_bawah:,.2f} m). Perbesar lebar tapak."
            )

        use_counterfort = st.checkbox("Gunakan Sirip Penahan (Counterfort)", value=False, key="5_c_cf")
        vol_sirip_total = 0.0
        luas_bekisting_sirip = 0.0

        if use_counterfort:
            col_cf1, col_cf2, col_cf3 = st.columns(3)
            t_bawah_sirip = col_cf1.number_input("Lebar Bawah Sirip (m)", value=float(max(l_heel, 0.0)),
                                                 min_value=0.0, key="5_c_tsb")
            t_atas_sirip = col_cf2.number_input("Lebar Atas Sirip (m)", value=0.0, min_value=0.0,
                                                help="Isi 0 untuk bentuk segitiga", key="5_c_tsa")
            t_tebal_sirip = col_cf3.number_input("Tebal Sirip (m)", value=0.3, min_value=0.0, key="5_c_tsirip")
            jarak_sirip = st.number_input("Jarak Antar Sirip (m)", value=2.5, min_value=0.1, key="5_c_jsirip")

            n_sirip = int(panjang / jarak_sirip) + 1
            vol_sirip_total = ((t_bawah_sirip + t_atas_sirip) / 2) * h * t_tebal_sirip * n_sirip
            sisi_miring_sirip = np.sqrt((t_bawah_sirip - t_atas_sirip) ** 2 + h ** 2)
            luas_bekisting_sirip = ((t_bawah_sirip + t_atas_sirip) * h) * n_sirip + (
                sisi_miring_sirip * t_tebal_sirip * n_sirip
            )
            st.caption(f"*{n_sirip} buah sirip · volume {vol_sirip_total:,.2f} m³*")

        vol_dinding = ((t_atas + t_bawah) / 2) * h * panjang
        vol_base = l_base * t_base * panjang
        vol_beton = vol_dinding + vol_base + vol_sirip_total

        h_galian_input = st.number_input("Kedalaman Galian (m)", value=float(t_base + 0.5),
                                         min_value=0.0, key="5_c_hgal_in")
        ws = st.number_input("Ruang Kerja Galian Total (m)", value=1.0, min_value=0.0,
                             help="Kelonggaran lebar galian total (0,5 m tiap sisi).", key="5_c_ws")
        vol_galian = (l_base + ws) * h_galian_input * panjang

        sisi_miring_dinding = np.sqrt(h ** 2 + (t_bawah - t_atas) ** 2)
        luas_bekisting = (h + sisi_miring_dinding) * panjang + (t_base * 2 * panjang) + luas_bekisting_sirip
        luas_muka_depan = h * panjang

        st.markdown("**Pekerjaan & AHSP**")
        if mode_proyek != "Bangunan Baru":
            p_bongkar = st.slider("Persen Bongkaran (%)", 0, 100, 100, key="5_c_sl_bongk")
            if st.checkbox("Pembongkaran Struktur Eksisting", value=True, key="5_c_cb_bongk"):
                tambah(f"Pembongkaran Struktur Eksisting ({p_bongkar}%)",
                       vol_beton * (p_bongkar / 100), "m³", harga_input("bongkar_m3", "5_c_bongkar"))

        if st.checkbox("Pekerjaan Galian (Termasuk Ruang Kerja)", value=True, key="5_c_cb_gal"):
            tambah("Pekerjaan Galian Struktur Tebing", vol_galian, "m³", harga_input("galian", "5_c_galian"))
        if st.checkbox("Pekerjaan Bekisting", value=True, key="5_c_cb_bek"):
            tambah("Pekerjaan Bekisting DPT", luas_bekisting, "m²", harga_input("bekisting", "5_c_bek"))
        if st.checkbox("Beton Struktur", value=True, key="5_c_cb_cor"):
            tambah(f"Beton DPT{' & Counterfort' if use_counterfort else ''}", vol_beton, "m³",
                   harga_input("beton_struktur", "5_c_beton"))
        if st.checkbox("Tulangan Utama D16-200", value=True, key="5_c_cb_besi"):
            r_besi = st.number_input("Rasio Besi (kg/m³)", value=150.0 if use_counterfort else 125.0,
                                     min_value=0.0, key="5_c_r_besi")
            tambah("Tulangan Utama Struktur DPT", vol_beton * r_besi, "kg", harga_input("besi", "5_c_besi"))

        st.markdown("**Material Timbunan**")
        jenis_timbunan = st.radio("Pilih Jenis Timbunan:", ["Tanah Kembali", "Sirtu / Material Pilihan"],
                                  horizontal=True, key="5_c_jtimb")
        if st.checkbox(f"Pekerjaan Urugan ({jenis_timbunan})", value=True, key="5_c_cb_timb"):
            kode_timb = "urugan_tanah" if jenis_timbunan == "Tanah Kembali" else "urugan_sirtu"
            h_timbunan = harga_input(kode_timb, f"5_c_timb_{kode_timb}")
            vol_ruang_timbunan = (max(l_heel, 0.0) * h * panjang) + (0.5 * (t_bawah - t_atas) * h * panjang)
            vol_timbunan_netto = max(0.0, vol_ruang_timbunan - vol_sirip_total)
            tambah(f"Pekerjaan Urugan {jenis_timbunan}", vol_timbunan_netto, "m³", h_timbunan)

        input_suling(luas_muka_depan, "5_c")

        # ---- Visualisasi ----
        fig, ax = plt.subplots(figsize=(6, 5))
        ax.add_patch(plt.Rectangle((0, -t_base), l_base, t_base, color="darkgray", ec="black"))
        ax.add_patch(plt.Polygon(
            [[l_toe, 0], [l_toe + t_bawah, 0], [l_toe + t_atas, h], [l_toe, h]],
            color="darkgray", ec="black"))
        if use_counterfort:
            ax.add_patch(plt.Polygon(
                [[l_toe + t_bawah, 0], [l_toe + t_bawah + t_bawah_sirip, 0],
                 [l_toe + t_atas + t_atas_sirip, h], [l_toe + t_atas, h]],
                color="gray", ec="black", alpha=0.5, label="Sirip Counterfort"))
        ax.add_patch(plt.Polygon(
            [[l_toe + t_bawah, 0], [l_base, 0], [l_base, h], [l_toe + t_atas, h]],
            color="saddlebrown", alpha=0.3,
            hatch="//" if jenis_timbunan != "Tanah Kembali" else "",
            label=f"Timbunan ({jenis_timbunan.split(' ')[0]})"))
        ax.text(l_base / 2, -t_base / 2, f"{l_base}m", ha="center", va="center", fontsize=9, color="white")
        ax.text(l_toe + t_atas / 2, h / 2, f"{h}m", ha="center", va="center",
                fontsize=9, color="white", rotation=90)
        ax.set_xlim(-1.0, l_base + 1.0)
        ax.set_ylim(-t_base - 1.0, h + 1)
        ax.set_aspect("equal")
        ax.set_xlabel("Lebar Struktur (m)")
        ax.set_ylabel("Tinggi/Elevasi (m)")
        ax.grid(True, linestyle="--", alpha=0.6)
        ax.legend(loc="upper right")

# =====================================================================
# LOGIKA 6. PONDASI BORE PILE
# =====================================================================
elif jenis_bangunan == "6. Pondasi Bore Pile":
    st.markdown("**Dimensi Bore Pile**")
    diameter = st.number_input("Diameter Pile (m)", value=0.6, min_value=0.0, key="6_d")
    kedalaman = st.number_input("Kedalaman Pile (m)", value=12.0, min_value=0.0, key="6_ked")
    jml_titik = st.number_input("Jumlah Titik", value=20, min_value=0, step=1, key="6_jml")

    area = np.pi * (diameter / 2) ** 2
    vol_total_beton = area * kedalaman * jml_titik
    st.info(f"Luas penampang {area:,.3f} m² · total beton **{vol_total_beton:,.2f} m³** "
            f"({jml_titik} titik × {kedalaman} m)")

    st.markdown("**Pekerjaan & AHSP**")
    if mode_proyek == "Rehabilitasi Struktur":
        p_bongkar = st.slider("Persen Titik Dibongkar (%)", 0, 100, 100, key="6_sl_bongk")
        if st.checkbox("Pembongkaran Struktur Eksisting", value=True, key="6_cb_bongk"):
            tambah(f"Pembongkaran Struktur Eksisting / Kepala Pile ({p_bongkar}%)",
                   jml_titik * (p_bongkar / 100), "Titik", harga_input("bongkar_titik", "6_bongkar"))

    if st.checkbox("Pekerjaan Pengeboran", value=True, key="6_cb_bor"):
        faktor_bor = st.number_input("Faktor Over-break Pengeboran", value=1.05, min_value=1.0,
                                     help="Kelebihan galian bor terhadap volume teoritis.", key="6_fbor")
        tambah("Pengeboran Bore Pile", vol_total_beton * faktor_bor, "m³", harga_input("bor_pile", "6_bor"))

    # PERBAIKAN v2: panjang casing adalah kedalaman pemasangan per titik,
    # bukan fungsi diameter. Casing umumnya hanya sedalam zona tanah lepas.
    if st.checkbox("Instalasi Temporary Casing", value=True, key="6_cb_cas"):
        panjang_casing = st.number_input(
            "Panjang Casing per Titik (m)", value=3.0, min_value=0.0,
            help="Kedalaman casing sementara, biasanya sepanjang zona tanah lepas di bagian atas.",
            key="6_pcas",
        )
        total_casing = panjang_casing * jml_titik
        st.caption(f"*{panjang_casing} m × {jml_titik} titik = **{total_casing:,.2f} m'***")
        tambah("Instalasi Temporary Casing", total_casing, "m'", harga_input("casing", "6_casing"))

    if st.checkbox("Beton Bore Pile", value=True, key="6_cb_cor"):
        faktor_beton = st.number_input("Faktor Pemborosan Beton", value=1.10, min_value=1.0,
                                       help="Kompensasi pembesaran lubang bor saat pengecoran.", key="6_fbeton")
        tambah("Beton Bore Pile", vol_total_beton * faktor_beton, "m³", harga_input("beton_struktur", "6_beton"))

    if st.checkbox("Tulangan Utama D16-200", value=True, key="6_cb_besi"):
        r_besi = st.number_input("Rasio Besi (kg/m³)", value=180.0, min_value=0.0, key="6_r_besi")
        tambah("Tulangan Utama D16-200 Bore Pile", vol_total_beton * r_besi, "kg", harga_input("besi", "6_besi"))

    fig, ax = plt.subplots(figsize=(5, 4))
    ax.add_patch(plt.Rectangle((-1, -kedalaman), 2, kedalaman, color="saddlebrown", alpha=0.12))
    ax.add_patch(plt.Rectangle((-diameter / 2, -kedalaman), diameter, kedalaman, color="gray", ec="black"))
    ax.plot([-1, 1], [0, 0], color="saddlebrown", lw=3)
    ax.set_xlim(-1, 1)
    ax.set_ylim(-kedalaman - 1, 1)
    ax.set_aspect("equal")
    ax.set_xlabel("Diameter (m)")
    ax.set_ylabel("Kedalaman (m)")
    ax.grid(True, linestyle="--", alpha=0.6)

# =====================================================================
# LOGIKA 7. PROTEKSI LERENG
# =====================================================================
elif jenis_bangunan == "7. Proteksi Lereng (Shotcrete & Soil Nailing)":
    st.markdown("**Dimensi Lereng/Tebing**")
    col_l1, col_l2 = st.columns(2)
    panjang = col_l1.number_input("Panjang Memanjang Lereng (m)", value=50.0, min_value=0.0, key="7_p")
    tinggi_miring = col_l2.number_input(
        "Panjang Miring Lereng (m)", value=15.0, min_value=0.0,
        help="Jarak dari kaki lereng ke puncak secara miring", key="7_tm",
    )
    luas_lereng = panjang * tinggi_miring
    st.info(f"**Total Luas Permukaan Lereng:** {luas_lereng:,.2f} m²")

    st.markdown("**Spesifikasi Shotcrete / Facing Beton**")
    col_s1, col_s2 = st.columns(2)
    t_bawah_shot = col_s1.number_input("Tebal Facing Bawah (m)", value=0.20, min_value=0.0, key="7_tbwh")
    t_atas_shot = col_s2.number_input("Tebal Facing Atas (m)", value=0.10, min_value=0.0,
                                      help="Jika 0, penampang menjadi segitiga.", key="7_tats")
    t_rata_rata = (t_bawah_shot + t_atas_shot) / 2
    lapis_wiremesh = st.number_input("Jumlah Lapis Wiremesh M10", value=1, min_value=1, step=1, key="7_wm")

    st.markdown("**Spesifikasi Soil Nailing**")
    pakai_nailing = st.checkbox("Gunakan Soil Nailing?", value=True, key="7_cb_nail")
    jml_nailing, panjang_nail = 0, 0.0
    if pakai_nailing:
        col_n1, col_n2 = st.columns(2)
        jarak_h = col_n1.number_input("Jarak Horizontal (m)", value=1.5, min_value=0.1, key="7_jh")
        jarak_v = col_n2.number_input("Jarak Vertikal (m)", value=1.5, min_value=0.1, key="7_jv")
        panjang_nail = st.number_input("Kedalaman Masuk Tanah (L) (m)", value=6.0, min_value=0.0, key="7_pn")
        jml_nailing = np.ceil(luas_lereng / (jarak_h * jarak_v)) if luas_lereng > 0 else 0
        st.caption(f"*Estimasi kebutuhan: **{int(jml_nailing)} titik** soil nailing*")

    st.markdown("**Pekerjaan & AHSP**")
    if st.checkbox("Kupas/Perapihan Permukaan Lereng", value=True, key="7_cb_kupas"):
        tambah("Perapihan & Pembersihan Permukaan Lereng", luas_lereng, "m²",
               harga_input("perapihan", "7_perapihan"))
    if st.checkbox("Shotcrete / Pengecoran Facing K-300", value=True, key="7_cb_shot"):
        st.caption(f"*Tebal rata-rata {t_rata_rata:,.3f} m × luas {luas_lereng:,.2f} m² = "
                   f"**{luas_lereng * t_rata_rata:,.2f} m³***")
        tambah("Pekerjaan Shotcrete / Facing Beton", luas_lereng * t_rata_rata, "m³",
               harga_input("shotcrete", "7_shot"))
    if st.checkbox(f"Pemasangan Wiremesh M10 ({lapis_wiremesh} Lapis)", value=True, key="7_cb_wm"):
        overlap = st.number_input("Faktor Overlap & Waste", value=1.10, min_value=1.0, key="7_ovl")
        tambah(f"Pemasangan Wiremesh M10 ({lapis_wiremesh} Lapis + Overlap)",
               luas_lereng * lapis_wiremesh * overlap, "m²", harga_input("wiremesh_m2", "7_wm"))
    if pakai_nailing and st.checkbox("Soil Nailing D25 Terpasang", value=True, key="7_cb_do_nail"):
        h_nailing = harga_input(
            "soil_nail", "7_nail",
            label="AHSP Soil Nailing (Rp/Titik) — sudah termasuk bor, besi D25, grouting, bearing plate & mur",
        )
        tambah(f"Soil Nailing D25 (Kedalaman {panjang_nail}m)", jml_nailing, "Titik", h_nailing)

    # ---- Visualisasi ----
    fig, ax = plt.subplots(figsize=(6, 5))
    sudut = np.radians(60)
    h_visual = tinggi_miring * np.sin(sudut)
    w_visual = tinggi_miring * np.cos(sudut)

    ax.add_patch(plt.Polygon(
        [[0, h_visual], [w_visual, 0], [w_visual + 10, 0],
         [w_visual + 10, h_visual + 10], [0, h_visual + 10]],
        color="saddlebrown", alpha=0.3, label="Tanah/Tebing Asli"))

    dx_atas, dy_atas = t_atas_shot * np.sin(sudut), t_atas_shot * np.cos(sudut)
    dx_bawah, dy_bawah = t_bawah_shot * np.sin(sudut), t_bawah_shot * np.cos(sudut)
    ax.add_patch(plt.Polygon(
        [[0, h_visual], [w_visual, 0], [w_visual - dx_bawah, -dy_bawah], [-dx_atas, h_visual - dy_atas]],
        color="gray", label=f"Facing (B:{t_bawah_shot}m, A:{t_atas_shot}m)"))

    if pakai_nailing:
        jarak_visual_v = tinggi_miring / 5 if tinggi_miring else 0
        for i in range(1, 5):
            L_tempuh = i * jarak_visual_v
            x_surf = L_tempuh * np.cos(sudut)
            y_surf = h_visual - (L_tempuh * np.sin(sudut))
            x_dalam = x_surf + (panjang_nail * np.sin(sudut))
            y_dalam = y_surf + (panjang_nail * np.cos(sudut))
            ax.plot([x_surf, x_dalam], [y_surf, y_dalam], color="black", lw=3)
            ax.plot([x_surf, x_dalam], [y_surf, y_dalam], color="red", lw=1.5, linestyle="--")
            ax.scatter([x_surf - (dx_atas + dx_bawah) / 4], [y_surf - (dy_atas + dy_bawah) / 4],
                       color="blue", s=80, zorder=5)
        ax.plot([], [], color="red", linestyle="--", label=f"Soil Nail D25 (L={panjang_nail}m)")
        ax.scatter([], [], color="blue", label="Bearing Plate")

    ax.set_xlim(-max(2.0, t_atas_shot + 1), w_visual + panjang_nail + 2)
    ax.set_ylim(-max(2.0, t_bawah_shot + 1), h_visual + panjang_nail)
    ax.set_aspect("equal")
    ax.set_title("Visualisasi Penampang Proteksi Lereng")
    ax.grid(True, linestyle="--", alpha=0.6)
    ax.legend(loc="lower right", fontsize=8)


# =====================================================================
# BLOK 2: REVIEW ESTIMASI SEMENTARA
# =====================================================================
st.markdown("---")
st.markdown("### 📝 Rincian Estimasi Sementara")
st.caption(f"**Kategori Saat Ini:** {kategori_pekerjaan}")
st.caption(
    "⚠️ *Ini hitungan sementara. Klik tombol **TAMBAHKAN KE MASTER REKAP** "
    "agar data tersimpan ke laporan final.*"
)

for pesan in peringatan:
    st.warning(pesan)

subtotal_now = 0.0
for item in item_to_add:
    biaya = item[1] * item[3]
    subtotal_now += biaya
    st.markdown(
        f"- **{item[0]}**<br>"
        f"<span style='color:gray; font-size:14px'>{item[1]:,.2f} {item[2]} × "
        f"Rp {item[3]:,.0f} = <b>Rp {biaya:,.0f}</b></span>",
        unsafe_allow_html=True,
    )

st.info(f"**Sub-Total Rincian Ini: Rp {subtotal_now:,.0f}**")

if len(item_to_add) > 0:
    if st.button("TAMBAHKAN KE MASTER REKAP", **LEBAR_PENUH, type="primary"):
        for item in item_to_add:
            vol = round(float(item[1]), 3)
            st.session_state.rekap_proyek.append({
                "Kategori": kategori_pekerjaan,
                "Pekerjaan": item[0],
                "Volume": vol,
                "Satuan": item[2],
                "AHSP": float(item[3]),
                "Total": vol * float(item[3]),  # konsisten dengan volume yang ditampilkan
            })
        st.success("Data berhasil ditambahkan ke tabel RAB di bawah.")

if fig is not None:
    st.markdown("---")
    st.pyplot(fig)
    plt.close(fig)


# =====================================================================
# FUNGSI PENYUSUN & EXPORT RAB
# =====================================================================
ROMAWI_MAP = [(1000, "M"), (900, "CM"), (500, "D"), (400, "CD"), (100, "C"), (90, "XC"),
              (50, "L"), (40, "XL"), (10, "X"), (9, "IX"), (5, "V"), (4, "IV"), (1, "I")]


def ke_romawi(n):
    hasil = ""
    for nilai, simbol in ROMAWI_MAP:
        while n >= nilai:
            hasil += simbol
            n -= nilai
    return hasil


def susun_rab(rekap):
    """
    Ubah daftar item mentah menjadi baris RAB berurut dengan penomoran
    hierarkis (I, I.1, I.2, ...) dan nilai sub-total per kategori.
    """
    df = pd.DataFrame(rekap)
    baris = []
    biaya_langsung = float(df["Total"].sum())
    kategori_urut = sorted(df["Kategori"].unique())

    for i, kat in enumerate(kategori_urut, start=1):
        df_kat = df[df["Kategori"] == kat]
        nama_kat = kat.split(". ", 1)[1] if ". " in kat else kat
        rom = ke_romawi(i)
        baris.append({"tipe": "header", "no": rom, "uraian": nama_kat.upper(),
                      "volume": None, "satuan": "", "harga": None, "jumlah": None})
        for j, (_, row) in enumerate(df_kat.iterrows(), start=1):
            baris.append({
                "tipe": "item", "no": f"{rom}.{j}", "uraian": row["Pekerjaan"],
                "volume": float(row["Volume"]), "satuan": row["Satuan"],
                "harga": float(row["AHSP"]), "jumlah": float(row["Total"]),
            })
        baris.append({"tipe": "subtotal", "no": "", "uraian": f"SUB-TOTAL {nama_kat.upper()}",
                      "volume": None, "satuan": "", "harga": None,
                      "jumlah": float(df_kat["Total"].sum())})
    return baris, biaya_langsung


def buat_excel(meta, baris, biaya_langsung, oh_pct, ppn_pct):
    """Susun file Excel RAB. Kolom Jumlah & rekap memakai FORMULA agar
    pengguna bisa mengubah volume/harga di Excel dan total ikut terhitung."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "RAB"

    tipis = Side(style="thin", color="000000")
    kotak = Border(left=tipis, right=tipis, top=tipis, bottom=tipis)
    arial = "Arial"
    fill_header = PatternFill("solid", fgColor="D9E1F2")
    fill_kat = PatternFill("solid", fgColor="F2F2F2")

    lebar = {"A": 8, "B": 52, "C": 12, "D": 10, "E": 18, "F": 20, "G": 10}
    for kol, w in lebar.items():
        ws.column_dimensions[kol].width = w

    # ---- Kop dokumen ----
    judul = [
        "PT PLN INDONESIA POWER",
        "UNIT BISNIS PEMBANGKITAN SAGULING - BIDANG PEMELIHARAAN SIPIL",
        "RENCANA ANGGARAN BIAYA (RAB)",
    ]
    for i, teks in enumerate(judul, start=1):
        ws.merge_cells(f"A{i}:G{i}")
        c = ws[f"A{i}"]
        c.value = teks
        c.font = Font(name=arial, size=13 if i == 1 else 11, bold=True)
        c.alignment = Alignment(horizontal="center")

    info = [
        ("Paket Pekerjaan", meta.get("paket", "")),
        ("Lokasi", meta.get("lokasi", "")),
        ("Tahun Anggaran", meta.get("tahun", "")),
        ("Nomor Dokumen", meta.get("nomor", "")),
    ]
    r = 5
    for label, nilai in info:
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(name=arial, size=10, bold=True)
        ws[f"C{r}"] = f": {nilai}"
        ws[f"C{r}"].font = Font(name=arial, size=10)
        r += 1

    # ---- Header tabel ----
    r += 1
    baris_header = r
    headers = ["NO", "URAIAN PEKERJAAN", "VOLUME", "SATUAN", "HARGA SATUAN (Rp)",
               "JUMLAH HARGA (Rp)", "BOBOT"]
    for idx, teks in enumerate(headers, start=1):
        c = ws.cell(row=r, column=idx, value=teks)
        c.font = Font(name=arial, size=10, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = kotak
        c.fill = fill_header
    r += 1

    baris_subtotal = []
    baris_pertama_data = r

    for b in baris:
        ws.cell(row=r, column=1, value=b["no"])
        ws.cell(row=r, column=2, value=b["uraian"])
        if b["tipe"] == "item":
            ws.cell(row=r, column=3, value=round(b["volume"], 3))
            ws.cell(row=r, column=4, value=b["satuan"])
            ws.cell(row=r, column=5, value=round(b["harga"], 2))
            ws.cell(row=r, column=6, value=f"=C{r}*E{r}")
        elif b["tipe"] == "subtotal":
            # cari rentang item di atasnya
            awal = r - 1
            while awal > baris_pertama_data and ws.cell(row=awal, column=6).value is not None \
                    and str(ws.cell(row=awal, column=6).value).startswith("="):
                awal -= 1
            ws.cell(row=r, column=6, value=f"=SUM(F{awal + 1}:F{r - 1})")
            baris_subtotal.append(r)

        for kol in range(1, 8):
            sel = ws.cell(row=r, column=kol)
            sel.border = kotak
            sel.font = Font(name=arial, size=10, bold=(b["tipe"] != "item"))
            if b["tipe"] in ("header", "subtotal"):
                sel.fill = fill_kat
            if kol in (3, 5, 6):
                sel.number_format = "#,##0.00" if kol == 3 else "#,##0"
                sel.alignment = Alignment(horizontal="right")
            elif kol in (1, 4, 7):
                sel.alignment = Alignment(horizontal="center")
            else:
                sel.alignment = Alignment(horizontal="left", wrap_text=True)
        r += 1

    baris_terakhir_data = r - 1

    # ---- Rekapitulasi biaya ----
    r += 1
    ref_subtotal = "+".join(f"F{x}" for x in baris_subtotal) if baris_subtotal else "0"
    baris_langsung = r
    ws[f"B{r}"] = "A. TOTAL BIAYA LANGSUNG"
    ws[f"F{r}"] = f"={ref_subtotal}"
    r += 1
    baris_oh = r
    ws[f"B{r}"] = "B. OVERHEAD & PROFIT"
    ws[f"E{r}"] = oh_pct / 100.0
    ws[f"F{r}"] = f"=F{baris_langsung}*E{r}"
    r += 1
    baris_ab = r
    ws[f"B{r}"] = "C. JUMLAH (A + B)"
    ws[f"F{r}"] = f"=F{baris_langsung}+F{baris_oh}"
    r += 1
    baris_ppn = r
    ws[f"B{r}"] = "D. PPN / PAJAK"
    ws[f"E{r}"] = ppn_pct / 100.0
    ws[f"F{r}"] = f"=F{baris_ab}*E{r}"
    r += 1
    baris_grand = r
    ws[f"B{r}"] = "GRAND TOTAL KONTRAK"
    ws[f"F{r}"] = f"=F{baris_ab}+F{baris_ppn}"

    for rr in [baris_langsung, baris_oh, baris_ab, baris_ppn, baris_grand]:
        for kol in range(1, 8):
            sel = ws.cell(row=rr, column=kol)
            sel.border = kotak
            sel.font = Font(name=arial, size=10, bold=True)
        ws.cell(row=rr, column=6).number_format = "#,##0"
        ws.cell(row=rr, column=5).number_format = "0.00%"
        ws.cell(row=rr, column=5).alignment = Alignment(horizontal="center")
    ws.cell(row=baris_grand, column=6).font = Font(name=arial, size=11, bold=True)

    # ---- Kolom bobot (persentase terhadap biaya langsung) ----
    for rr in range(baris_pertama_data, baris_terakhir_data + 1):
        if ws.cell(row=rr, column=6).value is not None:
            c = ws.cell(row=rr, column=7, value=f"=IFERROR(F{rr}/$F${baris_langsung},0)")
            c.number_format = "0.00%"
            c.border = kotak
            c.font = Font(name=arial, size=10)
            c.alignment = Alignment(horizontal="center")

    # ---- Catatan & tanda tangan ----
    r += 2
    ws[f"A{r}"] = ("Catatan: kolom Jumlah Harga dihitung dengan formula Volume × Harga Satuan. "
                   "Persentase Overhead dan PPN dapat diubah pada kolom Harga Satuan di baris B dan D.")
    ws[f"A{r}"].font = Font(name=arial, size=9, italic=True)
    r += 3
    ws[f"E{r}"] = f"{meta.get('lokasi', '')}, {datetime.date.today().strftime('%d %B %Y')}"
    ws[f"E{r}"].font = Font(name=arial, size=10)
    r += 1
    ws[f"E{r}"] = "Disusun oleh,"
    ws[f"E{r}"].font = Font(name=arial, size=10)
    r += 4
    ws[f"E{r}"] = meta.get("penyusun", "") or "(...........................)"
    ws[f"E{r}"].font = Font(name=arial, size=10, bold=True, underline="single")

    ws.freeze_panes = ws[f"A{baris_header + 1}"]

    # ---- Sheet rekapitulasi ----
    ws2 = wb.create_sheet("Rekapitulasi")
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 52
    ws2.column_dimensions["C"].width = 22
    ws2["A1"] = "REKAPITULASI BIAYA"
    ws2["A1"].font = Font(name=arial, size=12, bold=True)
    for idx, teks in enumerate(["NO", "URAIAN KATEGORI", "JUMLAH (Rp)"], start=1):
        c = ws2.cell(row=3, column=idx, value=teks)
        c.font = Font(name=arial, size=10, bold=True)
        c.border = kotak
        c.fill = fill_header
        c.alignment = Alignment(horizontal="center")
    rr = 4
    no_kat = 1
    for b, baris_src in zip([x for x in baris if x["tipe"] == "subtotal"], baris_subtotal):
        ws2.cell(row=rr, column=1, value=ke_romawi(no_kat))
        ws2.cell(row=rr, column=2, value=b["uraian"].replace("SUB-TOTAL ", ""))
        ws2.cell(row=rr, column=3, value=f"=RAB!F{baris_src}")
        for kol in range(1, 4):
            sel = ws2.cell(row=rr, column=kol)
            sel.border = kotak
            sel.font = Font(name=arial, size=10)
        ws2.cell(row=rr, column=3).number_format = "#,##0"
        rr += 1
        no_kat += 1
    ws2.cell(row=rr, column=2, value="GRAND TOTAL KONTRAK").font = Font(name=arial, size=10, bold=True)
    ws2.cell(row=rr, column=3, value=f"=RAB!F{baris_grand}").font = Font(name=arial, size=10, bold=True)
    ws2.cell(row=rr, column=3).number_format = "#,##0"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def buat_pdf(meta, baris, biaya_langsung, oh_pct, ppn_pct):
    """Susun RAB dalam bentuk PDF siap cetak (butuh paket reportlab)."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    st_judul = ParagraphStyle("judul", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=13)
    st_sub = ParagraphStyle("sub", parent=styles["Normal"], fontName="Helvetica", fontSize=9)
    st_sel = ParagraphStyle("sel", parent=styles["Normal"], fontName="Helvetica", fontSize=8, leading=10)
    st_sel_b = ParagraphStyle("selb", parent=st_sel, fontName="Helvetica-Bold")

    elemen = [
        Paragraph("PT PLN INDONESIA POWER - UBP SAGULING", st_judul),
        Paragraph("RENCANA ANGGARAN BIAYA (RAB) - BIDANG PEMELIHARAAN SIPIL", st_judul),
        Spacer(1, 6),
        Paragraph(f"<b>Paket Pekerjaan:</b> {meta.get('paket','-')}", st_sub),
        Paragraph(f"<b>Lokasi:</b> {meta.get('lokasi','-')} &nbsp;&nbsp; "
                  f"<b>Tahun Anggaran:</b> {meta.get('tahun','-')} &nbsp;&nbsp; "
                  f"<b>Nomor:</b> {meta.get('nomor','-')}", st_sub),
        Spacer(1, 8),
    ]

    data = [[Paragraph(x, st_sel_b) for x in
             ["NO", "URAIAN PEKERJAAN", "VOLUME", "SAT.", "HARGA SATUAN (Rp)", "JUMLAH (Rp)", "BOBOT"]]]
    gaya_baris = []
    for idx, b in enumerate(baris, start=1):
        tebal = b["tipe"] != "item"
        s = st_sel_b if tebal else st_sel
        vol = f"{b['volume']:,.2f}" if b["volume"] is not None else ""
        hrg = f"{b['harga']:,.0f}" if b["harga"] is not None else ""
        jml = f"{b['jumlah']:,.0f}" if b["jumlah"] is not None else ""
        bobot = (f"{b['jumlah'] / biaya_langsung * 100:,.2f}%"
                 if (b["tipe"] == "item" and biaya_langsung) else "")
        data.append([Paragraph(b["no"], s), Paragraph(b["uraian"], s), Paragraph(vol, s),
                     Paragraph(b["satuan"], s), Paragraph(hrg, s), Paragraph(jml, s),
                     Paragraph(bobot, s)])
        if tebal:
            gaya_baris.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#F2F2F2")))

    oh = biaya_langsung * oh_pct / 100.0
    ppn = (biaya_langsung + oh) * ppn_pct / 100.0
    total = biaya_langsung + oh + ppn
    rekap = [
        ("A. TOTAL BIAYA LANGSUNG", biaya_langsung),
        (f"B. OVERHEAD & PROFIT ({oh_pct}%)", oh),
        ("C. JUMLAH (A + B)", biaya_langsung + oh),
        (f"D. PPN / PAJAK ({ppn_pct}%)", ppn),
        ("GRAND TOTAL KONTRAK", total),
    ]
    for label, nilai in rekap:
        data.append([Paragraph("", st_sel_b), Paragraph(label, st_sel_b), Paragraph("", st_sel_b),
                     Paragraph("", st_sel_b), Paragraph("", st_sel_b),
                     Paragraph(f"{nilai:,.0f}", st_sel_b), Paragraph("", st_sel_b)])
        gaya_baris.append(("BACKGROUND", (0, len(data) - 1), (-1, len(data) - 1),
                           colors.HexColor("#E8EEF7")))

    tabel = Table(data, colWidths=[18 * mm, 100 * mm, 24 * mm, 16 * mm, 35 * mm, 40 * mm, 18 * mm],
                  repeatRows=1)
    tabel.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E1F2")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 1), (2, -1), "RIGHT"),
        ("ALIGN", (4, 1), (5, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (3, 0), (3, -1), "CENTER"),
        ("ALIGN", (6, 0), (6, -1), "CENTER"),
    ] + gaya_baris))

    elemen.append(tabel)
    elemen.append(Spacer(1, 14))
    elemen.append(Paragraph(
        f"{meta.get('lokasi','')}, {datetime.date.today().strftime('%d %B %Y')}<br/>Disusun oleh,"
        f"<br/><br/><br/><br/><b>{meta.get('penyusun','') or '(...........................)'}</b>",
        st_sub))
    doc.build(elemen)
    return buffer.getvalue()


# =====================================================================
# BLOK 3: LAPORAN RAB & MANAJEMEN DATA
# =====================================================================
st.divider()
st.markdown("### 📊 Laporan Rencana Anggaran Biaya")

if st.session_state.rekap_proyek:
    # ---------------- Edit / hapus item ----------------
    with st.expander("✏️ Edit / Hapus Item Tersimpan"):
        opsi_edit = [
            f"{i+1}. {item['Pekerjaan']} ({item['Kategori'].split('.')[0]})"
            for i, item in enumerate(st.session_state.rekap_proyek)
        ]
        pilihan_edit = st.selectbox("Pilih Item:", ["-- Pilih Item --"] + opsi_edit, key="select_edit")

        if pilihan_edit != "-- Pilih Item --":
            idx_edit = int(pilihan_edit.split(".")[0]) - 1
            item_terpilih = st.session_state.rekap_proyek[idx_edit]
            st.info(
                f"**Data Saat Ini:**\n- Vol: {item_terpilih['Volume']} {item_terpilih['Satuan']}"
                f"\n- AHSP: Rp {item_terpilih['AHSP']:,.0f}"
            )
            persen_adj = st.slider("Persentase Penyesuaian Volume (%)", 0, 200, 100, step=1,
                                   key=f"adj_{idx_edit}")
            vol_hitung = float(item_terpilih["Volume"]) * (persen_adj / 100.0)
            val_vol = st.number_input(f"Edit Volume Akhir ({item_terpilih['Satuan']})",
                                      value=float(vol_hitung), key=f"ev_{idx_edit}_{persen_adj}")
            val_ahsp = st.number_input("Edit AHSP Akhir (Rp)", value=float(item_terpilih["AHSP"]),
                                       key=f"ea_{idx_edit}")
            col_e1, col_e2 = st.columns(2)
            if col_e1.button("💾 Update", key=f"upd_{idx_edit}", **LEBAR_PENUH):
                st.session_state.rekap_proyek[idx_edit]["Volume"] = round(val_vol, 3)
                st.session_state.rekap_proyek[idx_edit]["AHSP"] = val_ahsp
                st.session_state.rekap_proyek[idx_edit]["Total"] = round(val_vol, 3) * val_ahsp
                st.rerun()
            if col_e2.button("🗑️ Hapus", key=f"del_{idx_edit}", **LEBAR_PENUH):
                st.session_state.rekap_proyek.pop(idx_edit)
                st.rerun()

    # ---------------- Draft proyek ----------------
    with st.expander("📁 Manajemen Draft Proyek (Simpan/Buka)"):
        uploaded_file = st.file_uploader("Buka Draft RAB (.json)", type="json")
        if uploaded_file is not None and st.button("📂 Muat File Draft Ini", **LEBAR_PENUH):
            try:
                draft_data = json.load(uploaded_file)
                if isinstance(draft_data, list):          # kompatibel dengan draft v1
                    st.session_state.rekap_proyek = draft_data
                else:
                    st.session_state.rekap_proyek = draft_data.get("items", [])
                    st.session_state.meta.update(draft_data.get("meta", {}))
                    if draft_data.get("ahsp"):
                        st.session_state.ahsp.update(draft_data["ahsp"])
                        st.session_state.ahsp_ver += 1
                st.success("Draft berhasil dimuat!")
                st.rerun()
            except Exception:
                st.error("File draft tidak valid atau rusak.")

        draft_json = json.dumps(
            {
                "versi": APP_VERSION,
                "meta": st.session_state.meta,
                "keuangan": {"overhead": overhead_pct, "ppn": ppn_pct},
                "ahsp": st.session_state.ahsp,
                "items": st.session_state.rekap_proyek,
            },
            indent=4,
        )
        st.download_button(
            "💾 Simpan Draft Saat Ini (.json)",
            data=draft_json,
            file_name="Draft_RAB_Pemeliharaan_Sipil.json",
            mime="application/json",
            **LEBAR_PENUH,
        )

    # ---------------- Tabel RAB ----------------
    baris_rab, biaya_langsung = susun_rab(st.session_state.rekap_proyek)

    tampil = []
    for b in baris_rab:
        tampil.append({
            "No": b["no"],
            "Uraian Pekerjaan": b["uraian"],
            "Volume": f"{b['volume']:,.2f} {b['satuan']}" if b["volume"] is not None else "",
            "Harga Satuan": f"Rp {b['harga']:,.0f}" if b["harga"] is not None else "",
            "Jumlah Harga": f"Rp {b['jumlah']:,.0f}" if b["jumlah"] is not None else "",
            "Bobot": (f"{b['jumlah'] / biaya_langsung * 100:,.2f}%"
                      if (b["tipe"] == "item" and biaya_langsung) else ""),
        })

    oh = biaya_langsung * (overhead_pct / 100)
    ppn = (biaya_langsung + oh) * (ppn_pct / 100)
    total_akhir = biaya_langsung + oh + ppn

    for label, nilai in [
        ("A. TOTAL BIAYA LANGSUNG", biaya_langsung),
        (f"B. OVERHEAD & PROFIT ({overhead_pct}%)", oh),
        ("C. JUMLAH (A + B)", biaya_langsung + oh),
        (f"D. PPN / PAJAK ({ppn_pct}%)", ppn),
        ("GRAND TOTAL KONTRAK", total_akhir),
    ]:
        tampil.append({"No": "", "Uraian Pekerjaan": label, "Volume": "",
                       "Harga Satuan": "", "Jumlah Harga": f"Rp {nilai:,.0f}", "Bobot": ""})

    st.dataframe(pd.DataFrame(tampil), **LEBAR_PENUH, hide_index=True)

    c_m1, c_m2, c_m3 = st.columns(3)
    c_m1.metric("Biaya Langsung", f"Rp {biaya_langsung/1e6:,.1f} jt")
    c_m2.metric("OH + PPN", f"Rp {(oh+ppn)/1e6:,.1f} jt")
    c_m3.metric("Grand Total", f"Rp {total_akhir/1e6:,.1f} jt")

    # ---------------- Export ----------------
    st.markdown("#### 📤 Export Dokumen")
    if not st.session_state.meta.get("paket"):
        st.caption("*Isi **Identitas Dokumen RAB** di atas agar kop dokumen terisi lengkap.*")

    nama_file = (st.session_state.meta.get("paket") or "RAB_Pemeliharaan_Sipil").replace(" ", "_")[:60]
    c_x1, c_x2 = st.columns(2)

    with c_x1:
        try:
            data_xlsx = buat_excel(st.session_state.meta, baris_rab, biaya_langsung,
                                   overhead_pct, ppn_pct)
            st.download_button("📊 Download Excel (.xlsx)", data=data_xlsx,
                               file_name=f"{nama_file}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               **LEBAR_PENUH)
        except ImportError:
            st.warning("Paket `openpyxl` belum terpasang. Jalankan: `pip install openpyxl`")
        except Exception as e:
            st.error(f"Gagal menyusun Excel: {e}")

    with c_x2:
        try:
            data_pdf = buat_pdf(st.session_state.meta, baris_rab, biaya_langsung,
                                overhead_pct, ppn_pct)
            st.download_button("📄 Download PDF (.pdf)", data=data_pdf,
                               file_name=f"{nama_file}.pdf", mime="application/pdf",
                               **LEBAR_PENUH)
        except ImportError:
            st.warning("Paket `reportlab` belum terpasang. Jalankan: `pip install reportlab`")
        except Exception as e:
            st.error(f"Gagal menyusun PDF: {e}")

    st.write("---")
    if st.button("🗑️ Kosongkan Master Rekap / Buat Proyek Baru", **LEBAR_PENUH):
        st.session_state.rekap_proyek = []
        st.rerun()
else:
    st.info("Tabel RAB masih kosong. Silakan tambah rincian estimasi di atas.")
